import io
import os
import re
import tempfile
import pandas as pd
from flask import Flask, request, send_file, Response, jsonify
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

"""GSTIN Match Checker backend for CA Automation"""

# ---------------- Normalization helpers ---------------- #

def normalize_invoice_number(invoice):
    if pd.isna(invoice):
        return ''
    inv = re.sub(r'/\d{2,4}(-\d{2,4})?$', '', str(invoice).strip().upper())
    parts = re.split(r'([/-])', inv)
    normalized_parts = []
    for part in parts:
        if part and part not in ['/', '-']:
            if part.isdigit():
                part = part.lstrip('0') or '0'
        normalized_parts.append(part)
    return ''.join(normalized_parts)


def unify_invoice_numbers(df, invoice_col='Invoice No.'):
    df['Normalized_Invoice'] = df[invoice_col].apply(normalize_invoice_number)
    invoice_map = {}
    if 'Forms' in df.columns:
        df_2b = df[df['Forms'] == '2B']
    else:
        df_2b = df.copy()
    for _, row in df_2b.iterrows():
        original = str(row[invoice_col])
        normalized = row['Normalized_Invoice']
        if '/' not in original:
            invoice_map[normalized] = original
    df[invoice_col] = df.apply(
        lambda row: invoice_map.get(row['Normalized_Invoice'], row[invoice_col]),
        axis=1
    )
    df.drop(columns='Normalized_Invoice', inplace=True)
    return df


def _normalize_header_token(s: str) -> str:
    if s is None:
        return ""
    s = str(s).upper().strip()
    for ch in [".", ",", "-", "_", "/", "\\", "(", ")", ":", ";"]:
        s = s.replace(ch, " ")
    s = " ".join(s.split())
    return s


def _canonical_header_map() -> dict:
    return {
        "FORMS": "Forms",
        "FORM": "Forms",
        "OPENING": "Forms",
        "GSTIN": "GSTIN NO.",
        "GSTIN NO": "GSTIN NO.",
        "GSTIN NUMBER": "GSTIN NO.",
        "GST NO": "GSTIN NO.",
        "INVOICE NO": "Invoice No.",
        "INVOICE NUMBER": "Invoice No.",
        "INV NO": "Invoice No.",
        "INV NUMBER": "Invoice No.",
        "INVOICE": "Invoice No.",
        "CGST": "CGST",
        "SGST": "SGST",
        "IGST": "IGST",
        "TAXABLE AMT": "Taxable Amt",
        "TAXABLE AMOUNT": "Taxable Amt",
        "TAXABLE VALUE": "Taxable Amt",
        "TAXABLE": "Taxable Amt",
    }


def _rename_headers_canonically(df: pd.DataFrame) -> pd.DataFrame:
    synonyms = _canonical_header_map()
    new_cols = []
    for col in df.columns:
        norm = _normalize_header_token(col)
        canonical = synonyms.get(norm)
        new_cols.append(canonical if canonical else str(col).strip())
    df = df.copy()
    df.columns = new_cols
    return df


def _find_sheet_with_headers(file_path):
    try:
        xls = pd.ExcelFile(file_path)
    except Exception:
        return None
    for s in xls.sheet_names:
        try:
            df_raw_probe = pd.read_excel(file_path, sheet_name=s, header=None)
        except Exception:
            continue
        for _, row in df_raw_probe.iterrows():
            row_values = row.fillna('').astype(str).tolist()
            norm_values = {_normalize_header_token(v) for v in row_values}
            if {"FORMS", "GSTIN NO"}.issubset(norm_values) or {"FORMS", "GSTIN"}.issubset(norm_values):
                return s
    return xls.sheet_names[0] if xls.sheet_names else None


def check_gstin_matching(file_path, sheet_name=None):
    ext = os.path.splitext(file_path)[1].lower()
    if ext != ".xlsx":
        raise ValueError("Unsupported file type. Please upload a .xlsx file.")
    if sheet_name is None:
        sheet_name = _find_sheet_with_headers(file_path)
        if sheet_name is None:
            raise ValueError("Could not detect a header row. Ensure your sheet has 'Forms' and 'GSTIN NO.' columns.")
    try:
        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    except Exception as e:
        raise ValueError(f"Failed to read Excel: {e}")
    header_row_idx = None
    for idx, row in df_raw.iterrows():
        row_values = row.fillna('').astype(str).tolist()
        norm_values = {_normalize_header_token(v) for v in row_values}
        if {"FORMS", "GSTIN NO"}.issubset(norm_values) or {"FORMS", "GSTIN"}.issubset(norm_values):
            header_row_idx = idx
            break
    if header_row_idx is None:
        raise ValueError("Could not find the header row. Make sure columns include 'Forms' and 'GSTIN NO.'")
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
    df.columns = df.columns.astype(str).str.strip()
    df = _rename_headers_canonically(df)
    required_cols = ['Forms', 'GSTIN NO.', 'Invoice No.', 'CGST', 'SGST', 'IGST']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Missing required columns: {missing_cols}. Please ensure your sheet has these columns.")
    df['Invoice No.'] = df['Invoice No.'].astype(str).str.strip().str.upper()
    df = unify_invoice_numbers(df)
    df_3b = df[df['Forms'].isin(['3B', 'Opening'])]
    df_2b = df[df['Forms'] == '2B']
    df_opening = df[df['Forms'] == 'Opening']
    gstin_invoice_3b = set(zip(df_3b['GSTIN NO.'], df_3b['Invoice No.']))
    gstin_invoice_2b = set(zip(df_2b['GSTIN NO.'], df_2b['Invoice No.']))
    gstin_invoice_opening = set(zip(df_opening['GSTIN NO.'], df_opening['Invoice No.']))
    df['Total GST'] = df['CGST'].fillna(0) + df['SGST'].fillna(0) + df['IGST'].fillna(0)
    gstin_invoice_totals = df.groupby(['GSTIN NO.', 'Invoice No.'])['Total GST'].sum().reset_index()
    gstin_invoice_totals.columns = ['GSTIN NO.', 'Invoice No.', 'Invoice Total GST']
    df = df.merge(gstin_invoice_totals, on=['GSTIN NO.', 'Invoice No.'], how='left')

    def get_final_remark(row):
        gstin = str(row['GSTIN NO.']).strip()
        invoice_no = str(row['Invoice No.']).strip()
        form = row['Forms']
        if pd.isna(gstin) or pd.isna(invoice_no):
            return 'Blank'
        if gstin == 'R364':
            return 'R364'
        in_3b = (gstin, invoice_no) in gstin_invoice_3b
        in_2b = (gstin, invoice_no) in gstin_invoice_2b
        if form != 'Opening' and gstin != 'R364':
            if in_3b and not in_2b:
                return 'NI2B'
            elif in_2b and not in_3b:
                return 'NIB'
        if gstin in df_3b['GSTIN NO.'].values and invoice_no not in df_2b[df_2b['GSTIN NO.'] == gstin]['Invoice No.'].values:
            return 'GSTIN Mismatch'
        elif gstin in df_2b['GSTIN NO.'].values and invoice_no not in df_3b[df_3b['GSTIN NO.'] == gstin]['Invoice No.'].values:
            return 'GSTIN Mismatch'
        if invoice_no in df_3b['Invoice No.'].values and gstin not in df_2b[df_2b['Invoice No.'] == invoice_no]['GSTIN NO.'].values:
            return 'Invoice Mismatch'
        elif invoice_no in df_2b['Invoice No.'].values and gstin not in df_3b[df_3b['Invoice No.'] == invoice_no]['GSTIN NO.'].values:
            return 'Invoice Mismatch'
        try:
            gst_total = float(row['Invoice Total GST'])
        except (ValueError, TypeError):
            return 'Mismatched'
        if -1 < gst_total < 1:
            if (gstin, invoice_no) in gstin_invoice_opening:
                return 'Opening Match'
            else:
                return 'Matched'
        return 'Mismatched'

    df['Remark'] = df.apply(get_final_remark, axis=1)
    df.drop(columns=['Total GST', 'Invoice Total GST'], inplace=True)
    df = df.sort_values(by=['Invoice No.', 'GSTIN NO.'])
    output_file = file_path.replace('.xlsx', '_checked.xlsx')
    df.to_excel(output_file, index=False)
    wb = load_workbook(output_file)
    ws = wb.active
    remark_col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Remark':
            remark_col_idx = idx
            break
    if remark_col_idx is not None:
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        for row in ws.iter_rows(min_row=2, min_col=remark_col_idx, max_col=remark_col_idx):
            for cell in row:
                if cell.value == 'Mismatched':
                    for entire_row_cell in ws[cell.row]:
                        entire_row_cell.fill = red_fill
    subtotal_columns = ['Taxable Amt', 'IGST', 'CGST', 'SGST']
    subtotal_values = {}
    header = [cell.value for cell in ws[1]]
    col_letter_map = {col: ws.cell(row=1, column=i+1).column_letter for i, col in enumerate(header)}
    for col in subtotal_columns:
        if col in header:
            col_letter = col_letter_map[col]
            max_row = ws.max_row
            subtotal_formula = f"=SUBTOTAL(9, {col_letter}2:{col_letter}{max_row})"
            subtotal_values[col_letter] = subtotal_formula
    ws.insert_rows(1)
    ws['A1'] = 'SUBTOTAL'
    for col_letter, formula in subtotal_values.items():
        ws[f"{col_letter}1"] = formula
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(output_file)
    return output_file


def _process_uploaded_excel(file_bytes, original_filename='uploaded.xlsx'):
    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1] or '.xlsx')
    try:
        tmp_in.write(file_bytes)
        tmp_in.flush()
        tmp_in.close()
        out_path = check_gstin_matching(tmp_in.name)
        if not out_path or not os.path.exists(out_path):
            raise RuntimeError('Processing failed: no output produced.')
        with open(out_path, 'rb') as f:
            output_bytes = f.read()
        download_name = os.path.basename(out_path)
        return download_name, output_bytes
    finally:
        try:
            os.remove(tmp_in.name)
        except Exception:
            pass
        try:
            if 'out_path' in locals() and out_path and os.path.exists(out_path):
                os.remove(out_path)
        except Exception:
            pass


def create_app():
    app = Flask(__name__)
    CORS(app, origins='*')

    @app.get('/health')
    def health():
        return jsonify({'status': 'ok'})

    @app.post('/api/gst/check')
    def gst_check():
        f = request.files.get('file')
        if not f or f.filename == '':
            return jsonify({'error': 'No file provided'}), 400
        try:
            download_name, output_bytes = _process_uploaded_excel(f.read(), f.filename)
        except Exception as e:
            return jsonify({'error': str(e)}), 400
        return send_file(
            io.BytesIO(output_bytes),
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return app


app = create_app()

if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    app.run(host='0.0.0.0', port=port)
